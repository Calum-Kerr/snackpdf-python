from flask import Blueprint, request, send_file, jsonify
import os
import tempfile
import logging
import fitz
import re

# Optional imports
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except (ImportError, ValueError) as e:
    PANDAS_AVAILABLE = False
    logging.warning(f"Pandas not available: {e}")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logging.basicConfig(level=logging.DEBUG)

pdf_to_excel_bp = Blueprint('pdf_to_excel', __name__)

STATIC_DIR = os.path.join(os.getcwd(), 'pdf_to_image', 'static')

# Ensure the static directory exists
if not os.path.exists(STATIC_DIR):
    os.makedirs(STATIC_DIR)

def extract_tables_from_pdf(pdf_path, extraction_mode='auto_detect', pages='all'):
    """Extract tables from PDF using PyMuPDF"""
    try:
        pdf = fitz.open(pdf_path)
        all_tables = []
        
        # Determine pages to process
        if pages == 'all':
            page_range = range(pdf.page_count)
        elif pages == 'first':
            page_range = range(1)
        else:
            # Handle page range like "1-5" or "2,4,6"
            if '-' in pages:
                start, end = map(int, pages.split('-'))
                page_range = range(start-1, min(end, pdf.page_count))
            else:
                page_nums = [int(p.strip())-1 for p in pages.split(',')]
                page_range = [p for p in page_nums if 0 <= p < pdf.page_count]
        
        for page_num in page_range:
            page = pdf.load_page(page_num)
            
            if extraction_mode == 'auto_detect':
                # Try to find tables using text positioning
                tables = extract_tables_by_position(page)
            elif extraction_mode == 'all_text':
                # Extract all text as simple rows
                tables = extract_all_text_as_table(page)
            elif extraction_mode == 'structured_data':
                # Look for structured data patterns
                tables = extract_structured_data(page)
            else:
                # Default to auto-detect
                tables = extract_tables_by_position(page)
            
            for table in tables:
                all_tables.append({
                    'page': page_num + 1,
                    'data': table
                })
        
        pdf.close()
        return all_tables
        
    except Exception as e:
        logging.error(f"Error extracting tables: {e}")
        raise

def extract_tables_by_position(page):
    """Extract tables by analyzing text positioning"""
    try:
        # Get text with position information
        text_dict = page.get_text("dict")
        
        # Group text by approximate rows based on y-coordinates
        rows = {}
        for block in text_dict["blocks"]:
            if "lines" in block:
                for line in block["lines"]:
                    y = round(line["bbox"][1], 1)  # Round y-coordinate
                    if y not in rows:
                        rows[y] = []
                    
                    for span in line["spans"]:
                        rows[y].append({
                            'text': span["text"].strip(),
                            'x': span["bbox"][0],
                            'font_size': span["size"]
                        })
        
        # Sort rows by y-coordinate
        sorted_rows = sorted(rows.items())
        
        # Detect table-like structures
        tables = []
        current_table = []
        
        for y, row_spans in sorted_rows:
            # Sort spans by x-coordinate
            row_spans.sort(key=lambda x: x['x'])
            
            # Check if this looks like a table row (multiple columns)
            if len(row_spans) >= 2:
                row_data = [span['text'] for span in row_spans if span['text']]
                if row_data:
                    current_table.append(row_data)
            else:
                # End of table if we have accumulated rows
                if len(current_table) >= 2:
                    tables.append(current_table)
                current_table = []
        
        # Add final table if exists
        if len(current_table) >= 2:
            tables.append(current_table)
        
        return tables
        
    except Exception as e:
        logging.error(f"Error in position-based extraction: {e}")
        return []

def extract_all_text_as_table(page):
    """Extract all text as a simple table structure"""
    try:
        text = page.get_text()
        lines = [line.strip() for line in text.split('\n') if line.strip()]
        
        # Simple approach: split lines by whitespace to create columns
        table_data = []
        for line in lines:
            # Split by multiple spaces or tabs
            columns = re.split(r'\s{2,}|\t', line)
            if len(columns) > 1:
                table_data.append(columns)
        
        return [table_data] if table_data else []
        
    except Exception as e:
        logging.error(f"Error in all-text extraction: {e}")
        return []

def extract_structured_data(page):
    """Extract structured data patterns"""
    try:
        text = page.get_text()
        
        # Look for common patterns like key-value pairs
        patterns = [
            r'([A-Za-z\s]+):\s*([^\n]+)',  # Key: Value
            r'([A-Za-z\s]+)\s+([0-9,.$%]+)',  # Label Number
        ]
        
        table_data = []
        for pattern in patterns:
            matches = re.findall(pattern, text)
            if matches:
                table_data.extend([[match[0].strip(), match[1].strip()] for match in matches])
        
        return [table_data] if table_data else []
        
    except Exception as e:
        logging.error(f"Error in structured data extraction: {e}")
        return []

def create_excel_file(tables, output_path, output_format='xlsx'):
    """Create Excel file from extracted tables"""
    try:
        if output_format.lower() == 'csv':
            # For CSV, create manually without pandas
            with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
                import csv
                writer = csv.writer(csvfile)
                for table_info in tables:
                    for row in table_info['data']:
                        writer.writerow(row)
                    writer.writerow([])  # Empty row between tables
        else:
            # For Excel formats, use openpyxl directly
            if not OPENPYXL_AVAILABLE:
                raise Exception("openpyxl not available for Excel output")

            wb = Workbook()
            # Remove default sheet
            wb.remove(wb.active)

            for i, table_info in enumerate(tables):
                sheet_name = f"Page_{table_info['page']}_Table_{i+1}"
                if len(sheet_name) > 31:  # Excel sheet name limit
                    sheet_name = f"P{table_info['page']}_T{i+1}"

                ws = wb.create_sheet(title=sheet_name)

                # Write data to worksheet
                for row_idx, row_data in enumerate(table_info['data'], 1):
                    for col_idx, cell_value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=str(cell_value))

            # If no sheets were created, create a default one
            if not wb.worksheets:
                ws = wb.create_sheet(title="No Data")
                ws.cell(row=1, column=1, value="No tables found in PDF")

            wb.save(output_path)

        logging.debug(f"Excel file created: {output_path}")

    except Exception as e:
        logging.error(f"Error creating Excel file: {e}")
        raise

@pdf_to_excel_bp.route('/pdf_to_excel', methods=['POST'])
def convert_pdf_to_excel():
    """Handle PDF to Excel conversion requests"""
    try:
        file = request.files['file']
        
        # Get options from form
        extraction_mode = request.form.get('extraction_mode', 'auto_detect')
        output_format = request.form.get('output_format', 'xlsx')
        pages = request.form.get('pages', 'all')
        page_range = request.form.get('page_range', '')
        
        # Use page_range if pages is set to 'range'
        if pages == 'range' and page_range:
            pages = page_range
        
        logging.debug(f"Converting PDF to Excel - Mode: {extraction_mode}, Format: {output_format}, Pages: {pages}")
        
        with tempfile.TemporaryDirectory() as temp_dir:
            # Save uploaded PDF
            pdf_path = os.path.join(temp_dir, file.filename)
            file.save(pdf_path)
            
            # Extract tables
            tables = extract_tables_from_pdf(pdf_path, extraction_mode, pages)
            
            if not tables:
                return jsonify({"error": "No tables or data found in the PDF"}), 400
            
            # Create output file
            base_name = os.path.splitext(file.filename)[0]
            output_filename = f"{base_name}.{output_format}"
            output_path = os.path.join(temp_dir, output_filename)
            
            create_excel_file(tables, output_path, output_format)
            
            # Determine MIME type
            mime_types = {
                'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'xls': 'application/vnd.ms-excel',
                'csv': 'text/csv'
            }
            
            return send_file(output_path, as_attachment=True,
                           download_name=output_filename,
                           mimetype=mime_types.get(output_format, 'application/octet-stream'))
            
    except Exception as e:
        logging.error(f"Error in PDF to Excel conversion: {e}")
        return jsonify({"error": str(e)}), 500
